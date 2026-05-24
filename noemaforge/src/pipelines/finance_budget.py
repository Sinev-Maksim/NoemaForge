#!/usr/bin/env python3
"""
=== NoemaForge File Header ===
File: noemaforge/src/pipelines/finance_budget.py
Zone: release/package
Version: 0.31.13.alpha-patched1
Created: 2026-05-14
Modified: 2026-05-14
Purpose: Manage NoemaForge pipeline catalog, runs, gates, artifacts and state.
Inputs: Command-line arguments, environment variables, package files and local NoemaForge runtime state as applicable.
Outputs: Structured command output, files, service state or UI state as documented by the caller.
Side effects: Limited to the documented NoemaForge paths, runtime state directories or systemd units used by this file.
Tests: Syntax validation plus the release setup selftest, consistency-audit and targeted smoke checks.
Notes: Code comments are English-only; user-facing localized text belongs in docs/i18n or locale JSON files.
=== End NoemaForge File Header ===
"""
from __future__ import annotations


# === NoemaForge Autodoc File Header ===
# File: src/pipelines/finance_budget.py
# Purpose: Implement the deterministic pipeline 'finance_budget'.
# Invoked by / imported from:
#   - src/noemaforge_core.py
# Public API / entry functions:
#   - init_db
#   - ingest
#   - generate_report
#   - run
#   - main
# Inputs:
#   - --inbox
#   - --day
#   - Common path inputs: /var/lib/noemaforge, /workspace/inbox/bank, /opt/noemaforge/configs/finance-categories.yaml, /opt/noemaforge/configs/clarifications.yaml
#   - Imports: __future__, csv, datetime, hashlib, json, os, re, shutil
# Output formats / side effects:
#   - SQLite databases
#   - JSON files
#   - copied filesystem artifacts
#   - YAML files
#   - Markdown files
# AutoDoc: refreshed 2026-04-09 (heuristic, review before trusting for policy work)
# === End NoemaForge Autodoc File Header ===

"""finance_budget.py (v0.11.0)

Deterministic bank exports -> ledger + monthly report pipeline.

Input:
- /workspace/inbox/bank/**
  Recommended: CSV exports.
  XLSX supported only if openpyxl is installed.

Output:
- Persistent ledger:
    /var/lib/noemaforge/routines/finance/ledger.sqlite
- Raw file snapshots (best-effort):
    /var/lib/noemaforge/routines/finance/raw/<sha256>.<ext>
- Monthly report:
    /var/lib/noemaforge/routines/finance/YYYY-MM/report.md
    /var/lib/noemaforge/routines/finance/YYYY-MM/summary.json
- Inquiries (clarifications for last N days):
    /var/lib/noemaforge/inquiries/open/*.json

This pipeline is spine code: no network, no LLM required.
"""


import csv
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import sqlite3
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import yaml
except Exception:
    yaml = None

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None  # type: ignore


BASE = "/var/lib/noemaforge"
FIN_DIR = os.path.join(BASE, "routines", "finance")
LEDGER_DB = os.path.join(FIN_DIR, "ledger.sqlite")
RAW_DIR = os.path.join(FIN_DIR, "raw")

INBOX_DEFAULT = "/workspace/inbox/bank"

INQUIRIES_DIR = os.path.join(BASE, "inquiries")
INQ_OPEN = os.path.join(INQUIRIES_DIR, "open")
INQ_CLOSED = os.path.join(INQUIRIES_DIR, "closed")
INQ_EXPIRED = os.path.join(INQUIRIES_DIR, "expired")

CFG_CATEGORIES = "/opt/noemaforge/configs/finance-categories.yaml"
CFG_CLARIFY = "/opt/noemaforge/configs/clarifications.yaml"


# === NoemaForge Autodoc Function Header ===
# Function: _nowz()
# Purpose: Implement the routine ' nowz'.
# Inputs:
#   - No explicit parameters.
# Called by:
#   - src/audit_remediation.py
#   - src/noemaforge_core.py
#   - src/bundles.py
#   - src/caps.py
#   - src/casebase.py
#   - src/coordinator_fanout.py
#   - src/dream_cycle.py
#   - src/fixture_bundle.py
# Calls:
#   - isoformat, utcnow
# Returns / emits: str
# === End NoemaForge Autodoc Function Header ===
def _nowz() -> str:
    return dt.datetime.utcnow().isoformat() + "Z"


# === NoemaForge Autodoc Function Header ===
# Function: _ensure_dir(p: str)
# Purpose: Implement the routine ' ensure dir'.
# Inputs:
#   - p: str
# Called by:
#   - src/casebase.py
#   - src/memory_system.py
#   - src/pipelines/photos_diary.py
#   - src/vstore.py
# Calls:
#   - makedirs
# Returns / emits: None
# Side effects:
#   - creates directories
# === End NoemaForge Autodoc Function Header ===
def _ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


# === NoemaForge Autodoc Function Header ===
# Function: _load_yaml(path: str)
# Purpose: Implement the routine ' load yaml'.
# Inputs:
#   - path: str
# Called by:
#   - src/brainctl.py
#   - src/noemaforge_core.py
#   - src/canary_runner.py
#   - src/daily_scheduler.py
#   - src/fixture_bundle.py
#   - src/flow_metrics.py
#   - src/incidents.py
#   - src/knowledge/policy.py
# Calls:
#   - open, safe_load
# Returns / emits: Dict[str, Any]
# Side effects:
#   - reads or writes files
# Key locals:
#   - f
# === End NoemaForge Autodoc Function Header ===
def _load_yaml(path: str) -> Dict[str, Any]:
    if yaml is None:
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception:
        return {}


# === NoemaForge Autodoc Function Header ===
# Function: _sha256_file(path: str)
# Purpose: Implement the routine ' sha256 file'.
# Inputs:
#   - path: str
# Called by:
#   - src/bootdoctor.py
#   - src/brainctl.py
#   - src/casebase.py
#   - src/doctor.py
#   - src/glove_agent.py
#   - src/localgw_uplink_agent.py
#   - src/model_registry.py
#   - src/prestart.py
# Calls:
#   - sha256, hexdigest, open, iter, update, read
# Returns / emits: str
# Side effects:
#   - reads or writes files
# Key locals:
#   - chunk, f, h
# === End NoemaForge Autodoc Function Header ===
def _sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


# === NoemaForge Autodoc Function Header ===
# Function: _connect()
# Purpose: Implement the routine ' connect'.
# Inputs:
#   - No explicit parameters.
# Called by:
#   - src/casebase.py
#   - src/dream_cycle.py
#   - src/knowledge/store.py
#   - src/memory_system.py
#   - src/roadmap.py
#   - src/task_tools.py
#   - src/taskqueue.py
#   - src/vstore.py
# Calls:
#   - _ensure_dir, connect
# Returns / emits: sqlite3.Connection
# Side effects:
#   - opens a database or socket connection
# Key locals:
#   - con
# === End NoemaForge Autodoc Function Header ===
def _connect() -> sqlite3.Connection:
    _ensure_dir(FIN_DIR)
    con = sqlite3.connect(LEDGER_DB)
    con.row_factory = sqlite3.Row
    return con


# === NoemaForge Autodoc Function Header ===
# Function: init_db()
# Purpose: Implement the routine 'init db'.
# Inputs:
#   - No explicit parameters.
# Called by:
#   - src/casebase.py
# Calls:
#   - _connect, cursor, execute, commit, close
# Returns / emits: None
# Side effects:
#   - opens a database or socket connection
#   - executes SQL or shell-like commands
# Key locals:
#   - con, cur
# === End NoemaForge Autodoc Function Header ===
def init_db() -> None:
    con = _connect()
    cur = con.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS files (
          file_sha256 TEXT PRIMARY KEY,
          rel_path TEXT,
          orig_path TEXT,
          ext TEXT,
          size INTEGER,
          ingested_at TEXT,
          parse_status TEXT
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS transactions (
          txn_id TEXT PRIMARY KEY,
          file_sha256 TEXT,
          txn_date TEXT,
          description TEXT,
          amount REAL,
          currency TEXT,
          category TEXT,
          category_confidence REAL,
          needs_clarification INTEGER,
          created_at TEXT,
          meta_json TEXT
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_txn_date ON transactions(txn_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_txn_category ON transactions(category)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_txn_file ON transactions(file_sha256)")

    con.commit()
    con.close()


# === NoemaForge Autodoc Function Header ===
# Function: _list_input_files(root: str)
# Purpose: Implement the routine ' list input files'.
# Inputs:
#   - root: str
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - walk, sort, isdir, lower, append, join, splitext
# Returns / emits: List[str]
# Side effects:
#   - appends to logs or files
# Key locals:
#   - ext, fn, out
# === End NoemaForge Autodoc Function Header ===
def _list_input_files(root: str) -> List[str]:
    out: List[str] = []
    if not os.path.isdir(root):
        return out
    for base, _dirs, files in os.walk(root):
        for fn in files:
            ext = os.path.splitext(fn)[1].lower()
            if ext in (".csv", ".tsv", ".txt", ".xlsx"):
                out.append(os.path.join(base, fn))
    out.sort()
    return out


# === NoemaForge Autodoc Function Header ===
# Function: _detect_delimiter(sample: str)
# Purpose: Implement the routine ' detect delimiter'.
# Inputs:
#   - sample: str
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - count
# Returns / emits: str
# Key locals:
#   - best, best_n, cands, d, n
# === End NoemaForge Autodoc Function Header ===
def _detect_delimiter(sample: str) -> str:
    # crude but works.
    cands = [";", ",", "\t", "|"]
    best = ","
    best_n = -1
    for d in cands:
        n = sample.count(d)
        if n > best_n:
            best_n = n
            best = d
    return best


# === NoemaForge Autodoc Function Header ===
# Function: _try_decode(raw: bytes)
# Purpose: Implement the routine ' try decode'.
# Inputs:
#   - raw: bytes
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - decode
# Returns / emits: str
# Key locals:
#   - enc
# === End NoemaForge Autodoc Function Header ===
def _try_decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("utf-8", "replace")


# === NoemaForge Autodoc Function Header ===
# Function: _parse_date(s: str)
# Purpose: Implement the routine ' parse date'.
# Inputs:
#   - s: str
# Called by:
#   - src/pipelines/photos_diary.py
# Calls:
#   - strip, fromisoformat, date, strptime
# Returns / emits: Optional[dt.date]
# Key locals:
#   - f, fmts, s
# === End NoemaForge Autodoc Function Header ===
def _parse_date(s: str) -> Optional[dt.date]:
    s = (s or "").strip()
    if not s:
        return None
    fmts = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]
    for f in fmts:
        try:
            return dt.datetime.strptime(s, f).date()
        except Exception:
            continue
    # try ISO-ish
    try:
        return dt.date.fromisoformat(s[:10])
    except Exception:
        return None


# === NoemaForge Autodoc Function Header ===
# Function: _parse_amount(s: str)
# Purpose: Implement the routine ' parse amount'.
# Inputs:
#   - s: str
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - strip, sub, startswith, endswith, replace, float, count
# Returns / emits: Optional[float]
# Key locals:
#   - neg, s, val
# === End NoemaForge Autodoc Function Header ===
def _parse_amount(s: str) -> Optional[float]:
    s = (s or "").strip()
    if not s:
        return None
    # strip currency symbols and spaces
    s = re.sub(r"[^0-9,\.\-\(\)]", "", s)
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    # handle comma decimal
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    # remove thousand separators like 1,234.56 or 1.234,56
    # if both present, last is decimal.
    if s.count(",") > 1 and s.count(".") == 0:
        s = s.replace(",", "")
    if s.count(".") > 1 and s.count(",") == 0:
        s = s.replace(".", "")

    try:
        val = float(s)
        return -val if neg else val
    except Exception:
        return None


# === NoemaForge Autodoc Function Header ===
# Function: _norm(s: str)
# Purpose: Implement the routine ' norm'.
# Inputs:
#   - s: str
# Called by:
#   - tools/checker/noemaforge_check.py
# Calls:
#   - sub, lower, strip
# Returns / emits: str
# === End NoemaForge Autodoc Function Header ===
def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


# === NoemaForge Autodoc Function Header ===
# Function: _find_col(headers: List[str], patterns: List[str])
# Purpose: Implement the routine ' find col'.
# Inputs:
#   - headers: List[str]
#   - patterns: List[str]
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - strip, lower, compile, enumerate, search
# Returns / emits: Optional[str]
# Key locals:
#   - hs, low, pat, rx
# === End NoemaForge Autodoc Function Header ===
def _find_col(headers: List[str], patterns: List[str]) -> Optional[str]:
    hs = [h.strip() for h in headers]
    low = [h.lower() for h in hs]
    for pat in patterns:
        rx = re.compile(pat, re.I)
        for i, h in enumerate(low):
            if rx.search(h):
                return hs[i]
    return None


# === NoemaForge Autodoc Function Header ===
# Function: _iter_rows_csv(path: str)
# Purpose: Implement the routine ' iter rows csv'.
# Inputs:
#   - path: str
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - read, _try_decode, _detect_delimiter, StringIO, DictReader, open, strip, items, str
# Returns / emits: Iterable[Dict[str, str]]
# Side effects:
#   - reads or writes files
# Key locals:
#   - delim, f, raw, raw_all, reader, row, text, txt
# === End NoemaForge Autodoc Function Header ===
def _iter_rows_csv(path: str) -> Iterable[Dict[str, str]]:
    raw = open(path, "rb").read(64 * 1024)
    txt = _try_decode(raw)
    delim = _detect_delimiter(txt)

    # Re-open full file
    raw_all = open(path, "rb").read()
    text = _try_decode(raw_all)

    # csv module expects file-like
    from io import StringIO

    f = StringIO(text)
    reader = csv.DictReader(f, delimiter=delim)
    for row in reader:
        # normalize keys
        yield {str(k or "").strip(): str(v or "").strip() for k, v in row.items()}


# === NoemaForge Autodoc Function Header ===
# Function: _iter_rows_xlsx(path: str)
# Purpose: Implement the routine ' iter rows xlsx'.
# Inputs:
#   - path: str
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - load_workbook, iter_rows, RuntimeError, strip, range, str, min, len
# Returns / emits: Iterable[Dict[str, str]]
# Key locals:
#   - header, r, row, rows, wb, ws
# === End NoemaForge Autodoc Function Header ===
def _iter_rows_xlsx(path: str) -> Iterable[Dict[str, str]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl_missing")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = None
    for r in rows:
        if header is None:
            header = [str(x or "").strip() for x in r]
            continue
        if not header:
            continue
        row = {header[i]: str(r[i] or "").strip() for i in range(min(len(header), len(r)))}
        yield row


# === NoemaForge Autodoc Function Header ===
# Function: _categorize(desc: str, rules: List[Dict[str, Any]])
# Purpose: Implement the routine ' categorize'.
# Inputs:
#   - desc: str
#   - rules: List[Dict[str, Any]]
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - _norm, min, strip, get, str, float, append, compile, search
# Returns / emits: Tuple[Optional[str], float, List[str]]
# Side effects:
#   - appends to logs or files
# Key locals:
#   - best_cat, best_score, cat, conf, hits, k, kw, matched, r, rx, rx_s, score
# === End NoemaForge Autodoc Function Header ===
def _categorize(desc: str, rules: List[Dict[str, Any]]) -> Tuple[Optional[str], float, List[str]]:
    text = _norm(desc)
    best_cat: Optional[str] = None
    best_score = 0.0
    matched: List[str] = []

    for r in rules:
        cat = str(r.get("category") or "").strip()
        if not cat:
            continue
        score = 0.0
        hits: List[str] = []

        for kw in (r.get("keywords") or []):
            k = _norm(str(kw))
            if k and k in text:
                score += float(r.get("weight") or 1.0)
                hits.append(k)

        for rx_s in (r.get("regex") or []):
            try:
                rx = re.compile(str(rx_s), re.I)
                if rx.search(desc or ""):
                    score += float(r.get("regex_weight") or 2.0)
                    hits.append(f"re:{rx_s}")
            except Exception:
                continue

        if score > best_score:
            best_score = score
            best_cat = cat
            matched = hits

    if best_cat is None:
        return None, 0.0, []

    # map score -> confidence (simple)
    conf = min(1.0, 0.35 + 0.15 * best_score)
    return best_cat, conf, matched


# === NoemaForge Autodoc Function Header ===
# Function: _clarify_window_days()
# Purpose: Implement the routine ' clarify window days'.
# Inputs:
#   - No explicit parameters.
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - _load_yaml, int, get
# Returns / emits: int
# Key locals:
#   - p, pol
# === End NoemaForge Autodoc Function Header ===
def _clarify_window_days() -> int:
    pol = _load_yaml(CFG_CLARIFY)
    p = (pol.get("policies") or {}).get("finance.transaction_category") or {}
    return int(p.get("ask_window_days") or 30)


# === NoemaForge Autodoc Function Header ===
# Function: _write_inquiry(txn: Dict[str, Any], stream_id: str = 'finance.budget')
# Purpose: Implement the routine ' write inquiry'.
# Inputs:
#   - txn: Dict[str, Any]
#   - stream_id: str = 'finance.budget'
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - _ensure_dir, join, exists, _clarify_window_days, isoformat, _nowz, open, dump, get, utcnow, timedelta
# Returns / emits: str
# Side effects:
#   - reads or writes files
#   - serializes structured data
# Key locals:
#   - days, deadline, f, inquiry_id, obj, path
# === End NoemaForge Autodoc Function Header ===
def _write_inquiry(txn: Dict[str, Any], *, stream_id: str = "finance.budget") -> str:
    _ensure_dir(INQ_OPEN)
    _ensure_dir(INQ_CLOSED)
    _ensure_dir(INQ_EXPIRED)

    inquiry_id = f"FIN-{txn['txn_id']}"
    path = os.path.join(INQ_OPEN, inquiry_id + ".json")
    if os.path.exists(path):
        return path

    days = _clarify_window_days()
    deadline = (dt.datetime.utcnow() + dt.timedelta(days=days)).isoformat() + "Z"

    obj = {
        "inquiry_id": inquiry_id,
        "created_at": _nowz(),
        "kind": "finance.transaction_category",
        "stream_id": stream_id,
        "project_id": "",
        "status": "open",
        "question": "Уточни категорию транзакции (пока не забылось):",
        "subject": f"{txn.get('txn_date')} | {txn.get('description')} | {txn.get('amount')} {txn.get('currency')}",
        "deadline_at": deadline,
        "related": [
            {
                "txn_id": txn.get("txn_id"),
                "txn_date": txn.get("txn_date"),
                "amount": txn.get("amount"),
                "currency": txn.get("currency"),
                "description": txn.get("description"),
                "source_file": txn.get("file_sha256"),
            }
        ],
        "provenance": {"pipeline": "finance_budget", "version": "0.10.1"},
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    return path


# === NoemaForge Autodoc Function Header ===
# Function: ingest(root: str)
# Purpose: Implement the routine 'ingest'.
# Inputs:
#   - root: str
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - init_db, _ensure_dir, _list_input_files, _load_yaml, _clarify_window_days, _connect, cursor, commit, close, isinstance, today, timedelta
# Returns / emits: Dict[str, Any]
# Side effects:
#   - opens a database or socket connection
# Key locals:
#   - amount, amt_col, cat, cat_cfg, con, conf, credit_col, cur, cur_col, currency, cv, d
# === End NoemaForge Autodoc Function Header ===
def ingest(root: str) -> Dict[str, Any]:
    init_db()
    _ensure_dir(RAW_DIR)

    files = _list_input_files(root)
    ingested = 0
    tx_added = 0
    tx_total = 0

    cat_cfg = _load_yaml(CFG_CATEGORIES)
    rules = (cat_cfg.get("rules") or []) if isinstance(cat_cfg.get("rules"), list) else []

    window_days = _clarify_window_days()
    window_start = dt.date.today() - dt.timedelta(days=window_days)

    con = _connect()
    cur = con.cursor()

    for path in files:
        try:
            sha = _sha256_file(path)
        except Exception:
            continue

        cur.execute("SELECT file_sha256 FROM files WHERE file_sha256=?", (sha,))
        if cur.fetchone():
            continue  # already ingested

        ext = os.path.splitext(path)[1].lower()
        st = os.stat(path)

        # snapshot raw file (best-effort)
        rel = f"{sha}{ext}"
        snap_path = os.path.join(RAW_DIR, rel)
        try:
            # avoid duplicate copy on races
            if not os.path.exists(snap_path):
                shutil.copy2(path, snap_path)
        except Exception:
            snap_path = ""

        cur.execute(
            "INSERT INTO files(file_sha256, rel_path, orig_path, ext, size, ingested_at, parse_status) VALUES(?,?,?,?,?,?,?)",
            (sha, rel, path, ext, int(st.st_size), _nowz(), "pending"),
        )
        ingested += 1

        parse_ok = True
        try:
            if ext == ".xlsx":
                rows = list(_iter_rows_xlsx(path))
            else:
                rows = list(_iter_rows_csv(path))
        except Exception:
            rows = []
            parse_ok = False

        if not rows:
            parse_ok = False

        if parse_ok:
            headers = list(rows[0].keys())
            date_col = _find_col(headers, [r"^date$", r"дата", r"data"]) or ""
            desc_col = _find_col(headers, [r"desc", r"опис", r"merchant", r"операц", r"детал"]) or ""
            amt_col = _find_col(headers, [r"amount", r"sum", r"сумм", r"valor", r"value"]) or ""
            debit_col = _find_col(headers, [r"debit", r"расход"]) or ""
            credit_col = _find_col(headers, [r"credit", r"приход"]) or ""
            cur_col = _find_col(headers, [r"currency", r"валют", r"moeda"]) or ""

            for row in rows:
                tx_total += 1
                d = _parse_date(row.get(date_col, "") if date_col else "")
                if d is None:
                    continue
                desc = row.get(desc_col, "") if desc_col else ""

                amount: Optional[float] = None
                if amt_col:
                    amount = _parse_amount(row.get(amt_col, ""))
                if amount is None and debit_col:
                    dv = _parse_amount(row.get(debit_col, ""))
                    if dv is not None:
                        amount = -abs(dv)
                if amount is None and credit_col:
                    cv = _parse_amount(row.get(credit_col, ""))
                    if cv is not None:
                        amount = abs(cv)
                if amount is None:
                    continue

                currency = (row.get(cur_col, "") if cur_col else "").strip() or "UNK"

                cat, conf, hits = _categorize(desc, rules)
                needs = 0

                if cat is None:
                    if d >= window_start:
                        needs = 1
                    else:
                        cat = "unknown"
                        conf = 0.0

                txn_key = f"{sha}|{d.isoformat()}|{desc}|{amount}|{currency}"
                txn_id = hashlib.sha256(txn_key.encode("utf-8")).hexdigest()[:24]

                meta = {"matched": hits, "source_path": path}

                cur.execute("SELECT txn_id FROM transactions WHERE txn_id=?", (txn_id,))
                if cur.fetchone():
                    continue

                cur.execute(
                    """
                    INSERT INTO transactions(txn_id, file_sha256, txn_date, description, amount, currency, category, category_confidence, needs_clarification, created_at, meta_json)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        txn_id,
                        sha,
                        d.isoformat(),
                        desc,
                        float(amount),
                        currency,
                        cat or "",
                        float(conf),
                        int(needs),
                        _nowz(),
                        json.dumps(meta, ensure_ascii=False),
                    ),
                )
                tx_added += 1

                if needs == 1:
                    _write_inquiry(
                        {
                            "txn_id": txn_id,
                            "txn_date": d.isoformat(),
                            "description": desc,
                            "amount": float(amount),
                            "currency": currency,
                            "file_sha256": sha,
                        }
                    )

        cur.execute("UPDATE files SET parse_status=? WHERE file_sha256=?", ("ok" if parse_ok else "failed", sha))

    con.commit()
    con.close()

    return {
        "ok": True,
        "input_root": root,
        "files_seen": len(files),
        "files_ingested": ingested,
        "tx_rows_seen": tx_total,
        "tx_added": tx_added,
    }


# === NoemaForge Autodoc Function Header ===
# Function: _month_str(day: dt.date)
# Purpose: Implement the routine ' month str'.
# Inputs:
#   - day: dt.date
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - strftime
# Returns / emits: str
# === End NoemaForge Autodoc Function Header ===
def _month_str(day: dt.date) -> str:
    return day.strftime("%Y-%m")


# === NoemaForge Autodoc Function Header ===
# Function: generate_report(for_day: dt.date)
# Purpose: Implement the routine 'generate report'.
# Inputs:
#   - for_day: dt.date
# Called by:
#   - No external Python callsite detected; may be internal-only, callback-based, or CLI-dispatched.
# Calls:
#   - init_db, _month_str, join, _ensure_dir, replace, _connect, cursor, execute, fetchall, close, append, sorted
# Returns / emits: Dict[str, Any]
# Side effects:
#   - opens a database or socket connection
#   - executes SQL or shell-like commands
#   - appends to logs or files
# Key locals:
#   - amt, by_cat, cat, con, cur, day_dir, end, expense, f, income, md_lines, month
# === End NoemaForge Autodoc Function Header ===
def generate_report(for_day: dt.date) -> Dict[str, Any]:
    init_db()

    month = _month_str(for_day)
    out_dir = os.path.join(FIN_DIR, month)
    _ensure_dir(out_dir)

    # fetch month txns
    start = for_day.replace(day=1)
    # next month
    if start.month == 12:
        end = dt.date(start.year + 1, 1, 1)
    else:
        end = dt.date(start.year, start.month + 1, 1)

    con = _connect()
    cur = con.cursor()
    cur.execute(
        "SELECT * FROM transactions WHERE txn_date>=? AND txn_date<? ORDER BY txn_date ASC",
        (start.isoformat(), end.isoformat()),
    )
    rows = cur.fetchall()
    con.close()

    income = 0.0
    expense = 0.0
    by_cat: Dict[str, float] = {}
    unknown: List[Dict[str, Any]] = []

    for r in rows:
        amt = float(r["amount"])
        cat = str(r["category"] or "unknown")
        if amt >= 0:
            income += amt
        else:
            expense += abs(amt)

        by_cat[cat] = by_cat.get(cat, 0.0) + (abs(amt) if amt < 0 else 0.0)

        if int(r["needs_clarification"] or 0) == 1:
            unknown.append(
                {
                    "txn_id": r["txn_id"],
                    "date": r["txn_date"],
                    "amount": amt,
                    "currency": r["currency"],
                    "description": r["description"],
                }
            )

    net = income - expense

    # write markdown
    md_lines: List[str] = []
    md_lines.append(f"# Домашний бюджет: {month}\n")
    md_lines.append(f"\nСгенерировано: {_nowz()}\n")
    md_lines.append(f"\nДоходы: **{income:.2f}**\n")
    md_lines.append(f"Расходы: **{expense:.2f}**\n")
    md_lines.append(f"Итог: **{net:.2f}**\n")

    md_lines.append("\n## Расходы по категориям\n")
    for cat, val in sorted(by_cat.items(), key=lambda x: x[1], reverse=True):
        if val <= 0:
            continue
        md_lines.append(f"- {cat}: {val:.2f}\n")

    md_lines.append("\n## Уточнения (последние дни)\n")
    if not unknown:
        md_lines.append("Нет транзакций, требующих уточнения.\n")
    else:
        md_lines.append(f"Нужно уточнить: {len(unknown)}\n")
        for u in unknown[:50]:
            md_lines.append(f"- {u['date']} | {u['amount']:.2f} {u['currency']} | {u['description']} (txn_id={u['txn_id']})\n")

    md_lines.append("\n## Проверка\n")
    md_lines.append("- Убедись, что выгрузки лежат в /workspace/inbox/bank (CSV предпочтительнее).\n")
    md_lines.append("- Проверь странные категории и пометь их вручную через inquiries (пока UI нет).\n")

    report_path = os.path.join(out_dir, "report.md")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("".join(md_lines))

    # Also write a per-day snapshot (required by recurring task contract).
    day_dir = os.path.join(FIN_DIR, for_day.strftime("%Y-%m-%d"))
    _ensure_dir(day_dir)
    report_day_path = os.path.join(day_dir, "report.md")
    try:
        with open(report_day_path, "w", encoding="utf-8") as f:
            f.write("".join(md_lines))
    except Exception:
        report_day_path = ""

    summary = {
        "kind": "finance.budget.summary",
        "month": month,
        "generated_at": _nowz(),
        "tx_count": len(rows),
        "income": income,
        "expense": expense,
        "net": net,
        "expense_by_category": by_cat,
        "needs_clarification": len(unknown),
    }
    summary_path = os.path.join(out_dir, "summary.json")
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return {
        "ok": True,
        "month": month,
        "report_path": report_path,
        "report_day_path": report_day_path,
        "summary_path": summary_path,
        "summary": summary,
    }


# === NoemaForge Autodoc Function Header ===
# Function: run(inbox: str, day: str)
# Purpose: Implement the routine 'run'.
# Inputs:
#   - inbox: str
#   - day: str
# Called by:
#   - src/bootdoctor.py
#   - src/brainctl.py
#   - src/firstboot_eval.py
#   - src/hwscan.py
#   - src/knowledge_maintainer.py
#   - src/lan_discovery.py
#   - src/localgateway.py
#   - src/localgw_connectors/ipp.py
# Calls:
#   - _ensure_dir, date, ingest, generate_report, strptime, get
# Returns / emits: Dict[str, Any]
# Key locals:
#   - ing, rep, target_day
# === End NoemaForge Autodoc Function Header ===
def run(*, inbox: str, day: str) -> Dict[str, Any]:
    _ensure_dir(FIN_DIR)
    _ensure_dir(RAW_DIR)
    _ensure_dir(INQ_OPEN)
    _ensure_dir(INQ_CLOSED)
    _ensure_dir(INQ_EXPIRED)

    target_day = dt.datetime.strptime(day, "%Y-%m-%d").date()

    ing = ingest(inbox)
    rep = generate_report(target_day)

    return {
        "ok": True,
        "day": day,
        "inbox": inbox,
        "ingest": ing,
        "report": rep,
        "summary": f"Бюджет {rep.get('month')}: доход {rep['summary']['income']:.2f}, расход {rep['summary']['expense']:.2f}, итог {rep['summary']['net']:.2f}.",
    }


# === NoemaForge Autodoc Function Header ===
# Function: main(argv: List[str])
# Purpose: Implement the routine 'main'.
# Inputs:
#   - argv: List[str]
# Called by:
#   - bootstrap/microvm/noemaforge-microvm-run.py
#   - src/bootdoctor.py
#   - src/brainctl.py
#   - src/noemaforge_core.py
#   - src/brainui.py
#   - src/canary_runner.py
#   - src/doctor.py
#   - src/dream_cycle.py
# Calls:
#   - ArgumentParser, add_argument, parse_args, run, print, dumps, get
# Returns / emits: int
# Side effects:
#   - serializes structured data
#   - spawns subprocesses or workers
# Key locals:
#   - ap, args, res
# === End NoemaForge Autodoc Function Header ===
def main(argv: List[str]) -> int:
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--inbox", default=INBOX_DEFAULT)
    ap.add_argument("--day", required=True)
    args = ap.parse_args(argv)

    res = run(inbox=args.inbox, day=args.day)
    print(json.dumps(res, ensure_ascii=False, indent=2))
    return 0 if res.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main(os.sys.argv[1:]))
